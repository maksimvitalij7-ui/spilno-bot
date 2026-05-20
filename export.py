"""
Экспорт данных в Excel с красивым форматированием
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

STATUS_LABELS = {
    "at_office":         "✅ В офисе",
    "on_the_way":        "🚗 В пути",
    "sick":              "🤒 Болеет",
    "day_off":           "🌴 Отгул",
    "remote":            "🏠 Удалённо",
    "at_work_remote_loc":"📍 Не в офисе",
    None:                "❓ Не отметился",
}

def generate_excel(db) -> str:
    data = db.get_export_data(days=30)
    wb = openpyxl.Workbook()

    # ── Лист 1: Детали ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Детали"

    # Заголовок
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Отчёт сотрудников — {datetime.now().strftime('%d.%m.%Y')}"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="2E4057")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Заголовки столбцов
    headers = ["Сотрудник", "Дата", "Статус", "Расстояние (м)", "Отчёт за день", "Отметка"]
    header_fill = PatternFill("solid", fgColor="4A90D9")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 22

    # Данные
    fills = {
        "at_office":         PatternFill("solid", fgColor="D4EDDA"),
        "on_the_way":        PatternFill("solid", fgColor="FFF3CD"),
        "sick":              PatternFill("solid", fgColor="F8D7DA"),
        "day_off":           PatternFill("solid", fgColor="F8D7DA"),
        "remote":            PatternFill("solid", fgColor="D1ECF1"),
        "at_work_remote_loc":PatternFill("solid", fgColor="FFF3CD"),
        None:                PatternFill("solid", fgColor="F5F5F5"),
    }

    for row_idx, row in enumerate(data, 3):
        status = row.get("checkin_status")
        fill = fills.get(status, fills[None])

        values = [
            row.get("name", ""),
            row.get("check_date", ""),
            STATUS_LABELS.get(status, "—"),
            int(row["distance_m"]) if row.get("distance_m") else "—",
            row.get("report_text", "") or "Не сдан",
            "✅" if row.get("report_text") else "❌",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left", vertical="center",
                wrap_text=(col_idx == 5)
            )
        ws.row_dimensions[row_idx].height = 18

    # Ширина столбцов
    col_widths = [25, 12, 20, 16, 50, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    # ── Лист 2: Сводка по сотрудникам ────────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")

    ws2.merge_cells("A1:E1")
    t = ws2["A1"]
    t.value = "Сводка по сотрудникам (последние 30 дней)"
    t.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="2E4057")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    heads2 = ["Сотрудник", "Дней в офисе", "Удалённо", "Пропусков", "Отчётов сдано"]
    for c, h in enumerate(heads2, 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    ws2.row_dimensions[2].height = 20

    # Агрегируем данные
    from collections import defaultdict
    summary = defaultdict(lambda: {"at_office": 0, "remote": 0, "absent": 0, "reports": 0, "name": ""})
    for row in data:
        uid = row.get("name", "?")
        summary[uid]["name"] = uid
        s = row.get("checkin_status")
        if s == "at_office":
            summary[uid]["at_office"] += 1
        elif s in ("remote", "at_work_remote_loc"):
            summary[uid]["remote"] += 1
        elif s in ("sick", "day_off"):
            summary[uid]["absent"] += 1
        if row.get("report_text"):
            summary[uid]["reports"] += 1

    for r_idx, (_, emp) in enumerate(sorted(summary.items()), 3):
        row_fill = PatternFill("solid", fgColor="F9F9F9") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [emp["name"], emp["at_office"], emp["remote"], emp["absent"], emp["reports"]]
        for c_idx, v in enumerate(vals, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

    for i, w in enumerate([28, 14, 14, 14, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = "A3"

    # Сохраняем
    os.makedirs("exports", exist_ok=True)
    filepath = f"exports/report_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    wb.save(filepath)
    return filepath
